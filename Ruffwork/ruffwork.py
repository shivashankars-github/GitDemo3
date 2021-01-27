import openpyxl

book = openpyxl.load_workbook("C:\\Users\\ssrk1\\Desktop\\New Microsoft Excel Worksheet.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=1)

list = []
for i in range(2, sheet.max_row+1):
    dict = {}
    for j in range (2, sheet.max_column+1):
        dict[sheet.cell(row=1, column=j).value]= sheet.cell(row=i,column=j).value
    list.append(dict)

print(list)
print("\n")
print("\n")
print("\n")

for i in range(1, sheet.max_row+1):
    for j in range(1,sheet.max_column+1):
        print(sheet.cell(row=i,column=j).value, end = "  ")
    print("\n")

