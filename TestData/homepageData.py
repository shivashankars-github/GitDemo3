import openpyxl
class HomePageData:

    #testdata = [{"name":"rahul", "email":"shetty", "gender":"Male" }, {"name":"shiva", "email":"shiva@infy", "gender":"Male" }]

    @staticmethod
    def get_test_data():
        book = openpyxl.load_workbook("C:\\Users\\ssrk1\\Desktop\\New Microsoft Excel Worksheet.xlsx")
        sheet = book.active
        list = []
        for i in range(2, sheet.max_row+1):
            dict = {}
            for x in range(2, sheet.max_column+1):
                dict[sheet.cell(row=1, column=x).value]=sheet.cell(row=i, column=x).value
            list.append(dict)
        return list