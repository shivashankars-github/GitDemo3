import openpyxl

book = openpyxl.load_workbook("C:\\Users\\ssrk1\\Desktop\\New Microsoft Excel Worksheet.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=1)

list = []
for i in range(2, sheet.max_row+1):
    dict = {}
    #if sheet.cell(row=i,column=1).value == "testcase2":
    for x in range(2, sheet.max_column+1):
        dict[sheet.cell(row=1, column=x).value]=sheet.cell(row=i, column=x).value
    list.append(dict)


print(list)




